#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sport-Analyse-Dashboard: Daten-Pipeline
Holt Spiele der naechsten 3 Tage + Statistiken + Quoten und berechnet Vorhersagen.

Quellen (alle kostenlos):
  - OpenLigaDB          -> 1./2./3. Bundesliga (Spielplaene + Ergebnisse)
  - ESPN (inoffiziell)  -> Premier League, Serie A, La Liga 1+2 (Spielplaene), ATP/WTA (Matches, Rankings)
  - football-data.co.uk -> historische Ergebnisse + Quoten (grosse Ligen)
  - TheSportsDB         -> Tennis-Fallback
  - The Odds API        -> optional, nur wenn ODDS_API_KEY gesetzt (3. Liga + Tennis-Quoten)

Ausgabe: data/data.json
"""

import csv
import io
import json
import math
import os
import re
import sys
import time
import unicodedata
from datetime import datetime, timedelta, timezone
from difflib import SequenceMatcher
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

BERLIN_UTC_OFFSET = None  # wird via zoneinfo bestimmt
try:
    from zoneinfo import ZoneInfo
    TZ = ZoneInfo("Europe/Berlin")
except Exception:
    TZ = timezone(timedelta(hours=2))

UA = {"User-Agent": "Mozilla/5.0 (SportDashboard/1.0; private hobby project)"}
DAYS_AHEAD = 3
ODDS_API_KEY = os.environ.get("ODDS_API_KEY", "").strip()

NOW = datetime.now(TZ)
TODAY = NOW.date()


# ----------------------------------------------------------------------------
# HTTP-Helfer
# ----------------------------------------------------------------------------

def http_get(url, retries=3, timeout=25, as_json=True):
    last_err = None
    for attempt in range(retries):
        try:
            req = Request(url, headers=UA)
            with urlopen(req, timeout=timeout) as resp:
                raw = resp.read()
            if as_json:
                return json.loads(raw.decode("utf-8", errors="replace"))
            return raw.decode("utf-8", errors="replace")
        except (URLError, HTTPError, json.JSONDecodeError, TimeoutError) as e:
            last_err = e
            time.sleep(1.5 * (attempt + 1))
    print(f"  WARN: {url} nicht erreichbar ({last_err})", file=sys.stderr)
    return None


# ----------------------------------------------------------------------------
# Team-Namen-Normalisierung (ESPN vs. football-data.co.uk vs. OpenLigaDB)
# ----------------------------------------------------------------------------

ALIASES = {
    # Premier League
    "manchester united": "man united", "manchester utd": "man united",
    "manchester city": "man city",
    "newcastle united": "newcastle", "newcastle utd": "newcastle",
    "tottenham hotspur": "tottenham", "spurs": "tottenham",
    "wolverhampton wanderers": "wolves", "wolverhampton": "wolves",
    "nottingham forest": "nott'm forest", "nottm forest": "nott'm forest",
    "west ham united": "west ham",
    "brighton & hove albion": "brighton", "brighton and hove albion": "brighton",
    "afc bournemouth": "bournemouth",
    "leeds united": "leeds",
    "leicester city": "leicester",
    "sheffield united": "sheffield utd",
    "afc sunderland": "sunderland",
    # Serie A
    "internazionale": "inter", "inter milan": "inter", "inter mailand": "inter",
    "ac milan": "milan", "ac mailand": "milan",
    "as roma": "roma",
    "ss lazio": "lazio",
    "ssc napoli": "napoli",
    "juventus turin": "juventus",
    "acf fiorentina": "fiorentina", "ac fiorentina": "fiorentina",
    "hellas verona": "verona",
    "us sassuolo": "sassuolo",
    "udinese calcio": "udinese",
    "atalanta bc": "atalanta", "atalanta bergamo": "atalanta",
    "bologna fc": "bologna",
    "torino fc": "torino",
    "genoa cfc": "genoa",
    "cagliari calcio": "cagliari",
    "us lecce": "lecce",
    "parma calcio": "parma",
    "como 1907": "como",
    "us cremonese": "cremonese",
    "delfino pescara": "pescara",
    # La Liga
    "atletico madrid": "ath madrid", "atlético madrid": "ath madrid", "atletico de madrid": "ath madrid",
    "athletic club": "ath bilbao", "athletic bilbao": "ath bilbao",
    "real sociedad": "sociedad",
    "real betis": "betis",
    "celta vigo": "celta", "rc celta": "celta", "celta de vigo": "celta",
    "rayo vallecano": "vallecano",
    "deportivo alaves": "alaves", "deportivo alavés": "alaves",
    "real valladolid": "valladolid",
    "rcd espanyol": "espanol", "espanyol": "espanol", "espanyol barcelona": "espanol",
    "rcd mallorca": "mallorca",
    "real oviedo": "oviedo",
    "ud las palmas": "las palmas",
    "cadiz cf": "cadiz",
    "getafe cf": "getafe",
    "girona fc": "girona",
    "sevilla fc": "sevilla",
    "valencia cf": "valencia",
    "villarreal cf": "villarreal",
    "elche cf": "elche",
    "levante ud": "levante",
    "ca osasuna": "osasuna",
    # La Liga 2 (häufige)
    "sporting gijon": "sp gijon", "sporting de gijon": "sp gijon",
    "deportivo la coruna": "la coruna", "deportivo de la coruna": "la coruna", "rc deportivo": "la coruna",
    "real zaragoza": "zaragoza",
    "cd tenerife": "tenerife",
    "ud almeria": "almeria",
    "racing santander": "santander", "real racing club": "santander",
    "cd leganes": "leganes",
    "granada cf": "granada",
    "cd castellon": "castellon",
    "burgos cf": "burgos",
    "cd mirandes": "mirandes",
    "sd eibar": "eibar",
    "sd huesca": "huesca",
    "cordoba cf": "cordoba",
    "malaga cf": "malaga",
    "albacete balompie": "albacete",
    "cadiz": "cadiz",
    "fc andorra": "andorra",
    "ceuta": "ceuta", "ad ceuta": "ceuta",
    "real sociedad b": "sociedad b",
    "cultural leonesa": "cul leonesa", "cultural y deportiva leonesa": "cul leonesa",
}


def norm_team(name):
    if not name:
        return ""
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    s = ALIASES.get(s, s)
    return s


def team_match(a, b):
    """Fuzzy-Vergleich zweier (bereits normalisierter) Teamnamen."""
    if a == b:
        return True
    if a and b and (a in b or b in a) and min(len(a), len(b)) >= 4:
        return True
    return SequenceMatcher(None, a, b).ratio() >= 0.82


# ----------------------------------------------------------------------------
# Fussball: Ligen-Konfiguration
# ----------------------------------------------------------------------------

LEAGUES = [
    # Internationale Wettbewerbe (Pokal-Modus: keine Tabellenplaetze)
    {"id": "wm", "name": "WM", "country": "INT", "flag": "\U0001F30D",
     "source": "espn", "espn": "fifa.world", "fdcuk": None, "cup": True},
    {"id": "em", "name": "EM", "country": "INT", "flag": "\U0001F1EA\U0001F1FA",
     "source": "espn", "espn": "uefa.euro", "fdcuk": None, "cup": True},
    {"id": "cl", "name": "Champions League", "country": "INT", "flag": "\U0001F3C6",
     "source": "espn", "espn": "uefa.champions", "fdcuk": None, "cup": True},
    {"id": "el", "name": "Europa League", "country": "INT", "flag": "\U0001F3C5",
     "source": "espn", "espn": "uefa.europa", "fdcuk": None, "cup": True},
    {"id": "ecl", "name": "Conference League", "country": "INT", "flag": "\U0001F396",
     "source": "espn", "espn": "uefa.europa.conf", "fdcuk": None, "cup": True},
    {"id": "bl1", "name": "1. Bundesliga", "country": "DE", "flag": "\U0001F1E9\U0001F1EA",
     "source": "openligadb", "oldb": "bl1", "fdcuk": "D1"},
    {"id": "bl2", "name": "2. Bundesliga", "country": "DE", "flag": "\U0001F1E9\U0001F1EA",
     "source": "openligadb", "oldb": "bl2", "fdcuk": "D2"},
    {"id": "bl3", "name": "3. Liga", "country": "DE", "flag": "\U0001F1E9\U0001F1EA",
     "source": "openligadb", "oldb": "bl3", "fdcuk": None},
    {"id": "pl", "name": "Premier League", "country": "EN", "flag": "\U0001F3F4\U000E0067\U000E0062\U000E0065\U000E006E\U000E0067\U000E007F",
     "source": "espn", "espn": "eng.1", "fdcuk": "E0"},
    {"id": "sa", "name": "Serie A", "country": "IT", "flag": "\U0001F1EE\U0001F1F9",
     "source": "espn", "espn": "ita.1", "fdcuk": "I1"},
    {"id": "ll1", "name": "La Liga", "country": "ES", "flag": "\U0001F1EA\U0001F1F8",
     "source": "espn", "espn": "esp.1", "fdcuk": "SP1"},
    {"id": "ll2", "name": "La Liga 2", "country": "ES", "flag": "\U0001F1EA\U0001F1F8",
     "source": "espn", "espn": "esp.2", "fdcuk": "SP2"},
    {"id": "fr1", "name": "Ligue 1", "country": "FR", "flag": "\U0001F1EB\U0001F1F7",
     "source": "espn", "espn": "fra.1", "fdcuk": "F1"},
    {"id": "fr2", "name": "Ligue 2", "country": "FR", "flag": "\U0001F1EB\U0001F1F7",
     "source": "espn", "espn": "fra.2", "fdcuk": "F2"},
    {"id": "nl1", "name": "Eredivisie", "country": "NL", "flag": "\U0001F1F3\U0001F1F1",
     "source": "espn", "espn": "ned.1", "fdcuk": "N1"},
    {"id": "nl2", "name": "Eerste Divisie", "country": "NL", "flag": "\U0001F1F3\U0001F1F1",
     "source": "espn", "espn": "ned.2", "fdcuk": None},
    {"id": "be1", "name": "Pro League (Belgien)", "country": "BE", "flag": "\U0001F1E7\U0001F1EA",
     "source": "espn", "espn": "bel.1", "fdcuk": "B1"},
    {"id": "be2", "name": "Challenger Pro League (Belgien)", "country": "BE", "flag": "\U0001F1E7\U0001F1EA",
     "source": "espn", "espn": "bel.2", "fdcuk": None},
]


def current_season_year():
    """Fussball-Saisonjahr: ab Juli -> aktuelles Jahr (Saison 2026/27 == 2026)."""
    return NOW.year if NOW.month >= 7 else NOW.year - 1


def fdcuk_season_code(year):
    """2026 -> '2627'"""
    return f"{str(year)[2:]}{str(year + 1)[2:]}"


# ----------------------------------------------------------------------------
# Fussball: Ergebnisse laden (fuer Form + Modell)
# ----------------------------------------------------------------------------

def load_results_openligadb(shortcut, years):
    """Alle beendeten Spiele der angegebenen Saisons."""
    results = []
    for y in years:
        data = http_get(f"https://api.openligadb.de/getmatchdata/{shortcut}/{y}")
        if not data:
            continue
        for m in data:
            if not m.get("matchIsFinished"):
                continue
            res = next((r for r in m.get("matchResults", [])
                        if r.get("resultTypeID") == 2), None)
            if res is None and m.get("matchResults"):
                res = m["matchResults"][-1]
            if res is None:
                continue
            try:
                dt = datetime.fromisoformat(m["matchDateTimeUTC"].replace("Z", "+00:00"))
            except Exception:
                continue
            results.append({
                "date": dt,
                "home": m["team1"]["teamName"], "away": m["team2"]["teamName"],
                "hg": res.get("pointsTeam1", 0), "ag": res.get("pointsTeam2", 0),
            })
    return results


def load_results_fdcuk(code, years):
    """Ergebnisse von football-data.co.uk CSVs."""
    results = []
    for y in years:
        url = f"https://www.football-data.co.uk/mmz4281/{fdcuk_season_code(y)}/{code}.csv"
        raw = http_get(url, as_json=False)
        if not raw:
            continue
        try:
            reader = csv.DictReader(io.StringIO(raw))
            for row in reader:
                d, hg, ag = row.get("Date"), row.get("FTHG"), row.get("FTAG")
                if not d or hg in (None, "") or ag in (None, ""):
                    continue
                try:
                    if len(d.split("/")[-1]) == 4:
                        dt = datetime.strptime(d, "%d/%m/%Y")
                    else:
                        dt = datetime.strptime(d, "%d/%m/%y")
                    dt = dt.replace(tzinfo=TZ)
                except Exception:
                    continue
                results.append({
                    "date": dt,
                    "home": row.get("HomeTeam", ""), "away": row.get("AwayTeam", ""),
                    "hg": int(float(hg)), "ag": int(float(ag)),
                    "odds": _extract_odds_row(row),
                })
        except Exception as e:
            print(f"  WARN: CSV {url}: {e}", file=sys.stderr)
    return results


def _extract_odds_row(row):
    for pre in ("Avg", "B365", "PS", "Max"):
        h, d, a = row.get(f"{pre}H"), row.get(f"{pre}D"), row.get(f"{pre}A")
        if h and d and a:
            try:
                return {"h": float(h), "d": float(d), "a": float(a), "src": pre}
            except ValueError:
                pass
    return None


def load_results_espn(slug, years):
    """Ergebnisse via ESPN-Scoreboard. ESPN akzeptiert nur begrenzte Zeitraeume,
    daher Abruf in 90-Tage-Bloecken."""
    results = []
    start = datetime(years[0], 7, 1)
    end = min(NOW.replace(tzinfo=None), datetime(years[-1] + 1, 6, 30))
    cur = start
    while cur < end:
        chunk_end = min(cur + timedelta(days=89), end)
        url = (f"https://site.api.espn.com/apis/site/v2/sports/soccer/{slug}/scoreboard"
               f"?dates={cur:%Y%m%d}-{chunk_end:%Y%m%d}&limit=1000")
        data = http_get(url, retries=2)
        cur = chunk_end + timedelta(days=1)
        if not data:
            continue
        for ev in data.get("events", []):
            try:
                comp = ev["competitions"][0]
                if comp.get("status", {}).get("type", {}).get("completed") is not True:
                    continue
                dt = datetime.fromisoformat(ev["date"].replace("Z", "+00:00"))
                home = next(c for c in comp["competitors"] if c["homeAway"] == "home")
                away = next(c for c in comp["competitors"] if c["homeAway"] == "away")
                results.append({
                    "date": dt,
                    "home": home["team"]["displayName"], "away": away["team"]["displayName"],
                    "hg": int(home.get("score", 0)), "ag": int(away.get("score", 0)),
                })
            except Exception:
                continue
    return results


COUNTRY_DE = {
    "Germany": "Deutschland", "Argentina": "Argentinien", "France": "Frankreich",
    "Spain": "Spanien", "Italy": "Italien", "Netherlands": "Niederlande",
    "Belgium": "Belgien", "Brazil": "Brasilien", "Switzerland": "Schweiz",
    "Austria": "Österreich", "Croatia": "Kroatien", "Denmark": "Dänemark",
    "Sweden": "Schweden", "Norway": "Norwegen", "Poland": "Polen",
    "Czechia": "Tschechien", "Czech Republic": "Tschechien",
    "Türkiye": "Türkei", "Turkey": "Türkei", "Greece": "Griechenland",
    "Hungary": "Ungarn", "Romania": "Rumänien", "Bulgaria": "Bulgarien",
    "Serbia": "Serbien", "Slovakia": "Slowakei", "Slovenia": "Slowenien",
    "Scotland": "Schottland", "Republic of Ireland": "Irland", "Ireland": "Irland",
    "Northern Ireland": "Nordirland", "Iceland": "Island", "Finland": "Finnland",
    "United States": "USA", "USA": "USA", "Mexico": "Mexiko", "Canada": "Kanada",
    "South Korea": "Südkorea", "Korea Republic": "Südkorea", "Japan": "Japan",
    "Australia": "Australien", "Morocco": "Marokko", "Tunisia": "Tunesien",
    "Algeria": "Algerien", "Egypt": "Ägypten", "Cameroon": "Kamerun",
    "Ivory Coast": "Elfenbeinküste", "Côte d'Ivoire": "Elfenbeinküste",
    "South Africa": "Südafrika", "Saudi Arabia": "Saudi-Arabien", "Qatar": "Katar",
    "Uzbekistan": "Usbekistan", "Jordan": "Jordanien", "Colombia": "Kolumbien",
    "Ecuador": "Ecuador", "Bolivia": "Bolivien", "Cape Verde": "Kap Verde",
    "New Zealand": "Neuseeland", "Albania": "Albanien", "Georgia": "Georgien",
    "Jamaica": "Jamaika", "Curacao": "Curaçao", "Russia": "Russland",
    "Ukraine": "Ukraine", "Portugal": "Portugal", "England": "England",
    "Wales": "Wales", "Uruguay": "Uruguay", "Chile": "Chile", "Peru": "Peru",
    "Paraguay": "Paraguay", "Venezuela": "Venezuela", "Panama": "Panama",
    "Costa Rica": "Costa Rica", "Honduras": "Honduras", "Haiti": "Haiti",
    "Senegal": "Senegal", "Ghana": "Ghana", "Nigeria": "Nigeria", "Iran": "Iran",
}


def de_team(name):
    return COUNTRY_DE.get(name, name)


# ----------------------------------------------------------------------------
# Fussball: kommende Spiele (naechste 3 Tage) + naechster Spieltag
# ----------------------------------------------------------------------------

def upcoming_openligadb(shortcut, year):
    data = http_get(f"https://api.openligadb.de/getmatchdata/{shortcut}/{year}")
    if not data:
        return [], None
    upcoming, first_future = [], None
    for m in data:
        if m.get("matchIsFinished"):
            continue
        try:
            dt = datetime.fromisoformat(m["matchDateTimeUTC"].replace("Z", "+00:00")).astimezone(TZ)
        except Exception:
            continue
        if dt < NOW - timedelta(hours=3):
            continue
        entry = {"dt": dt, "home": m["team1"]["teamName"], "away": m["team2"]["teamName"],
                 "homeIcon": m["team1"].get("teamIconUrl"), "awayIcon": m["team2"].get("teamIconUrl"),
                 "matchday": (m.get("group") or {}).get("groupOrderID")}
        if first_future is None or dt < first_future["dt"]:
            first_future = entry
        if TODAY <= dt.date() <= TODAY + timedelta(days=DAYS_AHEAD - 1):
            upcoming.append(entry)
    return upcoming, first_future


def upcoming_espn(slug):
    upcoming = []
    seen = set()
    for i in range(DAYS_AHEAD):
        d = TODAY + timedelta(days=i)
        data = http_get(f"https://site.api.espn.com/apis/site/v2/sports/soccer/{slug}/scoreboard?dates={d:%Y%m%d}")
        if not data:
            continue
        for ev in data.get("events", []):
            try:
                comp = ev["competitions"][0]
                state = comp.get("status", {}).get("type", {}).get("state")
                if state not in ("pre",):
                    continue
                dt = datetime.fromisoformat(ev["date"].replace("Z", "+00:00")).astimezone(TZ)
                home = next(c for c in comp["competitors"] if c["homeAway"] == "home")
                away = next(c for c in comp["competitors"] if c["homeAway"] == "away")
                k = (home["team"]["displayName"], away["team"]["displayName"], dt.date().isoformat())
                if k in seen:
                    continue
                seen.add(k)
                if not (TODAY <= dt.date() <= TODAY + timedelta(days=DAYS_AHEAD - 1)):
                    continue
                upcoming.append({
                    "dt": dt,
                    "home": home["team"]["displayName"], "away": away["team"]["displayName"],
                    "homeIcon": (home["team"].get("logos") or [{}])[0].get("href") if home["team"].get("logos") else home["team"].get("logo"),
                    "awayIcon": (away["team"].get("logos") or [{}])[0].get("href") if away["team"].get("logos") else away["team"].get("logo"),
                    "matchday": None,
                })
            except Exception:
                continue
    # naechstes Spiel ausserhalb des Fensters suchen (fuer "Saisonpause"-Anzeige)
    first_future = None
    if not upcoming:
        end = TODAY + timedelta(days=120)
        data = http_get(f"https://site.api.espn.com/apis/site/v2/sports/soccer/{slug}/scoreboard"
                        f"?dates={TODAY:%Y%m%d}-{end:%Y%m%d}&limit=50")
        if data:
            for ev in data.get("events", []):
                try:
                    comp = ev["competitions"][0]
                    if comp.get("status", {}).get("type", {}).get("state") != "pre":
                        continue
                    dt = datetime.fromisoformat(ev["date"].replace("Z", "+00:00")).astimezone(TZ)
                    home = next(c for c in comp["competitors"] if c["homeAway"] == "home")
                    away = next(c for c in comp["competitors"] if c["homeAway"] == "away")
                    e = {"dt": dt, "home": home["team"]["displayName"],
                         "away": away["team"]["displayName"], "matchday": None}
                    if first_future is None or dt < first_future["dt"]:
                        first_future = e
                except Exception:
                    continue
    return upcoming, first_future


# ----------------------------------------------------------------------------
# Vorhersage-Modell (Poisson mit Zeitabklingen)
# ----------------------------------------------------------------------------

DECAY_HALF_LIFE_DAYS = 240.0


def _weight(days_ago):
    return 0.5 ** (max(days_ago, 0) / DECAY_HALF_LIFE_DAYS)


def build_strengths(results):
    """Angriffs-/Abwehrstaerken je Team (normalisiert), plus Heimvorteil."""
    if not results:
        return {}, 1.4, 1.15
    tw = {}   # team -> [w_sum, gf_w, ga_w]
    total_w = total_goals_w = 0.0
    home_goals_w = away_goals_w = 0.0
    for r in results:
        days = (NOW - (r["date"] if r["date"].tzinfo else r["date"].replace(tzinfo=TZ))).days
        w = _weight(days)
        h, a = norm_team(r["home"]), norm_team(r["away"])
        for team, gf, ga in ((h, r["hg"], r["ag"]), (a, r["ag"], r["hg"])):
            t = tw.setdefault(team, [0.0, 0.0, 0.0])
            t[0] += w
            t[1] += w * gf
            t[2] += w * ga
        total_w += w
        total_goals_w += w * (r["hg"] + r["ag"])
        home_goals_w += w * r["hg"]
        away_goals_w += w * r["ag"]
    league_avg = (total_goals_w / (2 * total_w)) if total_w else 1.4
    home_adv = (home_goals_w / away_goals_w) ** 0.5 if away_goals_w > 0 else 1.15
    strengths = {}
    for team, (w, gf, ga) in tw.items():
        if w < 2:   # zu wenig Daten
            strengths[team] = (1.0, 1.0, w)
            continue
        att = (gf / w) / league_avg
        dfn = (ga / w) / league_avg
        strengths[team] = (att, dfn, w)
    return strengths, league_avg, home_adv


def find_team(strengths, name):
    n = norm_team(name)
    if n in strengths:
        return strengths[n]
    for key, val in strengths.items():
        if team_match(n, key):
            return val
    return None


def poisson_probs(mu_h, mu_a, max_goals=9):
    ph = [math.exp(-mu_h) * mu_h ** k / math.factorial(k) for k in range(max_goals + 1)]
    pa = [math.exp(-mu_a) * mu_a ** k / math.factorial(k) for k in range(max_goals + 1)]
    p_home = p_draw = p_away = p_over25 = p_btts = 0.0
    best_score, best_p = (0, 0), -1.0
    for i in range(max_goals + 1):
        for j in range(max_goals + 1):
            p = ph[i] * pa[j]
            if i > j:
                p_home += p
            elif i == j:
                p_draw += p
            else:
                p_away += p
            if i + j >= 3:
                p_over25 += p
            if i > 0 and j > 0:
                p_btts += p
            if p > best_p:
                best_p, best_score = p, (i, j)
    return p_home, p_draw, p_away, p_over25, best_score, p_btts


def predict_match(strengths, league_avg, home_adv, home, away):
    sh, sa = find_team(strengths, home), find_team(strengths, away)
    if sh is None or sa is None:
        return None
    att_h, def_h, wh = sh
    att_a, def_a, wa = sa
    mu_h = max(0.15, league_avg * att_h * def_a * home_adv)
    mu_a = max(0.1, league_avg * att_a * def_h / home_adv)
    p_h, p_d, p_a, p_o25, score, p_btts = poisson_probs(mu_h, mu_a)
    conf = "hoch" if min(wh, wa) >= 8 else ("mittel" if min(wh, wa) >= 4 else "niedrig")
    return {
        "pHome": round(p_h, 4), "pDraw": round(p_d, 4), "pAway": round(p_a, 4),
        "pOver25": round(p_o25, 4), "pBtts": round(p_btts, 4),
        "xgHome": round(mu_h, 2), "xgAway": round(mu_a, 2),
        "tipScore": f"{score[0]}:{score[1]}",
        "confidence": conf,
    }


# ----------------------------------------------------------------------------
# Form, Tabelle, H2H
# ----------------------------------------------------------------------------

def team_form(results, team, n=5):
    n_t = norm_team(team)
    played = []
    for r in sorted(results, key=lambda x: x["date"]):
        h, a = norm_team(r["home"]), norm_team(r["away"])
        if team_match(n_t, h):
            played.append("S" if r["hg"] > r["ag"] else ("U" if r["hg"] == r["ag"] else "N"))
        elif team_match(n_t, a):
            played.append("S" if r["ag"] > r["hg"] else ("U" if r["hg"] == r["ag"] else "N"))
    return played[-n:]


def compute_table(results, season_start):
    table = {}
    for r in results:
        if r["date"].replace(tzinfo=r["date"].tzinfo or TZ) < season_start:
            continue
        h, a = norm_team(r["home"]), norm_team(r["away"])
        for t in (h, a):
            table.setdefault(t, {"pts": 0, "gd": 0, "gp": 0, "name": None})
        table[h]["name"] = table[h]["name"] or r["home"]
        table[a]["name"] = table[a]["name"] or r["away"]
        table[h]["gp"] += 1
        table[a]["gp"] += 1
        table[h]["gd"] += r["hg"] - r["ag"]
        table[a]["gd"] += r["ag"] - r["hg"]
        if r["hg"] > r["ag"]:
            table[h]["pts"] += 3
        elif r["hg"] < r["ag"]:
            table[a]["pts"] += 3
        else:
            table[h]["pts"] += 1
            table[a]["pts"] += 1
    ranked = sorted(table.items(), key=lambda kv: (-kv[1]["pts"], -kv[1]["gd"]))
    return {team: i + 1 for i, (team, _) in enumerate(ranked)}, len(ranked)


def table_pos(table, team):
    n = norm_team(team)
    if n in table:
        return table[n]
    for k, v in table.items():
        if team_match(n, k):
            return v
    return None


def head_to_head(results, home, away, n=5):
    nh, na = norm_team(home), norm_team(away)
    meetings = []
    for r in sorted(results, key=lambda x: x["date"], reverse=True):
        rh, ra = norm_team(r["home"]), norm_team(r["away"])
        if (team_match(nh, rh) and team_match(na, ra)) or (team_match(nh, ra) and team_match(na, rh)):
            meetings.append({
                "date": r["date"].strftime("%d.%m.%Y"),
                "home": r["home"], "away": r["away"],
                "score": f"{r['hg']}:{r['ag']}",
            })
        if len(meetings) >= n:
            break
    return meetings


# ----------------------------------------------------------------------------
# Quoten (football-data.co.uk fixtures.csv + optional The Odds API)
# ----------------------------------------------------------------------------

def load_fixture_odds_fdcuk():
    """fixtures.csv enthaelt Quoten fuer anstehende Spiele der grossen Ligen."""
    raw = http_get("https://www.football-data.co.uk/fixtures.csv", as_json=False)
    odds = {}
    if not raw:
        return odds
    try:
        reader = csv.DictReader(io.StringIO(raw))
        for row in reader:
            div = row.get("Div", "")
            o = _extract_odds_row(row)
            if not o:
                continue
            key = (div, norm_team(row.get("HomeTeam", "")), norm_team(row.get("AwayTeam", "")))
            odds[key] = o
    except Exception as e:
        print(f"  WARN fixtures.csv: {e}", file=sys.stderr)
    return odds


def find_odds(odds_map, div, home, away):
    if not div:
        return None
    nh, na = norm_team(home), norm_team(away)
    direct = odds_map.get((div, nh, na))
    if direct:
        return direct
    for (d, h, a), o in odds_map.items():
        if d == div and team_match(nh, h) and team_match(na, a):
            return o
    return None


def odds_api_get(sport_key):
    if not ODDS_API_KEY:
        return []
    url = (f"https://api.the-odds-api.com/v4/sports/{sport_key}/odds/"
           f"?apiKey={ODDS_API_KEY}&regions=eu&markets=h2h&oddsFormat=decimal")
    return http_get(url) or []


def odds_api_football(sport_key, home, away):
    """Sucht 1X2-Quoten fuer ein Spiel bei The Odds API (falls Key gesetzt)."""
    events = odds_api_cache.get(sport_key)
    if events is None:
        events = odds_api_get(sport_key)
        odds_api_cache[sport_key] = events
    nh, na = norm_team(home), norm_team(away)
    for ev in events:
        eh, ea = norm_team(ev.get("home_team", "")), norm_team(ev.get("away_team", ""))
        if not (team_match(nh, eh) and team_match(na, ea)):
            continue
        for bm in ev.get("bookmakers", []):
            for market in bm.get("markets", []):
                if market.get("key") != "h2h":
                    continue
                out = {o.get("name"): o.get("price") for o in market.get("outcomes", [])}
                h = out.get(ev.get("home_team"))
                a = out.get(ev.get("away_team"))
                d = out.get("Draw")
                if h and a:
                    return {"h": h, "d": d, "a": a, "src": bm.get("title", "OddsAPI")}
    return None


odds_api_cache = {}

ODDS_API_SPORT_KEYS = {
    "bl1": "soccer_germany_bundesliga",
    "bl2": "soccer_germany_bundesliga2",
    "bl3": "soccer_germany_liga3",
    "pl": "soccer_epl",
    "sa": "soccer_italy_serie_a",
    "ll1": "soccer_spain_la_liga",
    "ll2": "soccer_spain_segunda_division",
    "fr1": "soccer_france_ligue_one",
    "fr2": "soccer_france_ligue_two",
    "nl1": "soccer_netherlands_eredivisie",
    "be1": "soccer_belgium_pro_league",
    "cl": "soccer_uefa_champs_league",
    "el": "soccer_uefa_europa_league",
    "ecl": "soccer_uefa_europa_conference_league",
    "wm": "soccer_fifa_world_cup",
    "em": "soccer_uefa_european_championship",
}


def implied_probs(o):
    """Normalisierte implizite Wahrscheinlichkeiten aus 1X2-Quoten."""
    try:
        inv = [1.0 / o["h"], 1.0 / o["d"], 1.0 / o["a"]]
    except (TypeError, ZeroDivisionError, KeyError):
        return None
    s = sum(inv)
    return [x / s for x in inv]


# ----------------------------------------------------------------------------
# Tennis (ESPN + TheSportsDB)
# ----------------------------------------------------------------------------

def espn_tennis_day(tour, day):
    """Angesetzte Matches eines Tages, tour in {atp, wta}."""
    data = http_get(f"https://site.api.espn.com/apis/site/v2/sports/tennis/{tour}/scoreboard?dates={day:%Y%m%d}")
    matches = []
    if not data:
        return matches
    for ev in data.get("events", []):
        tour_name = ev.get("name", "")
        loc = (ev.get("locations") or [{}])
        venue = ""
        try:
            venue = ev.get("circuit", {}).get("fullName", "") or ""
        except Exception:
            pass
        groupings = ev.get("groupings") or []
        comps = []
        for g in groupings:
            gname = (g.get("grouping") or {}).get("displayName", "")
            gl = gname.lower()
            if "doubles" in gl:
                continue
            # Kombinierte Turniere: der WTA-Feed enthaelt teils auch Herren-Matches
            # (und umgekehrt). Nur die zum Tour-Kontext passenden Singles uebernehmen.
            if "women" in gl:
                gender = "W"
            elif "men" in gl:
                gender = "M"
            else:
                gender = None
            if (tour == "atp" and gender == "W") or (tour == "wta" and gender == "M"):
                continue
            comps.extend([(c, gname) for c in g.get("competitions", [])])
        if not groupings:
            comps = [(c, "") for c in ev.get("competitions", [])]
        for comp, gname in comps:
            try:
                st = comp.get("status", {}).get("type", {})
                if st.get("state") != "pre":
                    continue
                dt = datetime.fromisoformat(comp["date"].replace("Z", "+00:00")).astimezone(TZ)
                if dt.date() != day:
                    continue
                players = []
                for c in comp.get("competitors", []):
                    ath = c.get("athlete", {}) or {}
                    # Athlete-ID steckt je nach Feed-Variante in id, uid ("a:1234")
                    # oder einem $ref-Link
                    aid = str(ath.get("id") or "")
                    if not aid:
                        blob = " ".join(str(x) for x in (
                            ath.get("uid"), c.get("uid"), ath.get("$ref"),
                            (ath.get("links") or [{}])[0].get("href") if ath.get("links") else ""))
                        m_id = re.search(r"a:(\d+)", blob) or re.search(r"/athletes/(\d+)", blob)
                        aid = m_id.group(1) if m_id else ""
                    players.append({
                        "name": ath.get("displayName") or ath.get("shortName") or "?",
                        "id": aid,
                    })
                if len(players) != 2:
                    continue
                # Platzhalter (Gegner noch offen) ueberspringen
                if any(p["name"].strip().upper() in ("TBD", "?", "") for p in players):
                    continue
                matches.append({
                    "dt": dt, "tour": tour.upper(), "tournament": tour_name,
                    "round": comp.get("round", {}).get("displayName", "") if isinstance(comp.get("round"), dict) else (comp.get("note") or gname),
                    "p1": players[0], "p2": players[1], "venue": venue,
                    "src": "espn",
                    "timeTBD": comp.get("timeValid") is False,
                })
            except Exception:
                continue
    return matches


def tsdb_tennis_day(day):
    data = http_get(f"https://www.thesportsdb.com/api/v1/json/123/eventsday.php?d={day:%Y-%m-%d}&s=Tennis")
    matches = []
    if not data or not data.get("events"):
        return matches
    for ev in data["events"]:
        league = ev.get("strLeague", "")
        tour = "ATP" if "ATP" in league.upper() else ("WTA" if "WTA" in league.upper() else None)
        if not tour:
            continue
        name = ev.get("strEvent", "")
        m = re.search(r"(.+?)\s+(\S+(?:\s\S+)?)\s+vs\s+(.+)$", name)
        if not m:
            continue
        tournament, p1, p2 = m.group(1).strip(), m.group(2).strip(), m.group(3).strip()
        t = ev.get("strTime") or "00:00:00"
        try:
            dt = datetime.strptime(f"{ev['dateEvent']} {t}", "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc).astimezone(TZ)
        except Exception:
            continue
        matches.append({
            "dt": dt, "tour": tour, "tournament": tournament, "round": "",
            "p1": {"name": p1, "id": "", "seed": None},
            "p2": {"name": p2, "id": "", "seed": None},
            "venue": "",
            "src": "tsdb",
        })
    return matches


def load_sackmann(tour, years_back=4):
    """Historische Matchdaten von tennis-data.co.uk (ZIPs mit Excel/CSV) fuer
    Belag-Bilanz, H2H und Rank-Fallback. Die ZIPs laedt der Workflow-Schritt
    nach /tmp/tennisdata/; hier gibt es zusaetzlich einen HTTP-Fallback."""
    rows = []
    base_dir = os.environ.get("TENNISDATA_DIR", "/tmp/tennisdata")
    for y in range(NOW.year - years_back, NOW.year + 1):
        path = os.path.join(base_dir, f"{tour}_{y}.xlsx")
        blob = None
        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    blob = f.read()
            except OSError:
                blob = None
        if blob is None:
            suffix = "" if tour == "atp" else "w"
            url = f"https://www.tennis-data.co.uk/{y}{suffix}/{y}.xlsx"
            blob = http_get_bytes(url)
        if not blob:
            continue
        try:
            for r in _read_xlsx_rows(blob):
                rows.append(_tennisdata_row(r))
        except Exception as e:
            print(f"  WARN tennis-data {tour} {y}: {e}", file=sys.stderr)
    return [r for r in rows if r]


def http_get_bytes(url, retries=2, timeout=40):
    for attempt in range(retries):
        try:
            req = Request(url, headers=UA)
            with urlopen(req, timeout=timeout) as resp:
                return resp.read()
        except Exception:
            time.sleep(1.5 * (attempt + 1))
    print(f"  WARN: {url} nicht erreichbar (binaer)", file=sys.stderr)
    return None


def _read_xlsx_rows(blob):
    try:
        import openpyxl
    except ImportError:
        print("  WARN: openpyxl fehlt - xlsx uebersprungen", file=sys.stderr)
        return []
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(rows_iter, [])]
    out = []
    for vals in rows_iter:
        out.append({header[i]: vals[i] for i in range(min(len(header), len(vals)))})
    wb.close()
    return out


def _tennisdata_row(r):
    """tennis-data.co.uk-Zeile -> internes Format."""
    w, l = r.get("Winner"), r.get("Loser")
    if not w or not l:
        return None
    d = r.get("Date")
    if hasattr(d, "strftime"):
        date = d.strftime("%Y%m%d")
    else:
        date = ""
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                date = datetime.strptime(str(d).split()[0], fmt).strftime("%Y%m%d")
                break
            except (ValueError, AttributeError):
                continue
    def _score():
        sets = []
        for i in range(1, 6):
            a, b = r.get(f"W{i}"), r.get(f"L{i}")
            if a in (None, "") or b in (None, ""):
                break
            try:
                sets.append(f"{int(float(a))}:{int(float(b))}")
            except (ValueError, TypeError):
                break
        return " ".join(sets)
    def _rank(key):
        v = r.get(key)
        try:
            return str(int(float(v))) if v not in (None, "", "NR") else ""
        except (ValueError, TypeError):
            return ""
    return {
        "date": date,
        "tourney": str(r.get("Tournament") or ""),
        "loc": str(r.get("Location") or ""),
        "surface": str(r.get("Surface") or "").strip().capitalize(),
        "w": player_key(str(w)),
        "l": player_key(str(l)),
        "wname": str(w),
        "lname": str(l),
        "wrank": _rank("WRank"),
        "lrank": _rank("LRank"),
        "score": _score(),
        "rnd": str(r.get("Round") or ""),
    }


def player_key(name):
    """Einheitlicher Schluessel fuer beide Namensformate:
    'Jannik Sinner' (ESPN) und 'Sinner J.' (tennis-data) -> 'sinner j'."""
    s = unicodedata.normalize("NFKD", name or "").encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z. ]", " ", s).strip()
    if not s:
        return ""
    toks = s.split()
    if len(toks) == 1:
        return toks[0].lower()
    # tennis-data-Format: Nachname(n) + Initial(en) mit Punkt am Ende
    if toks[-1].endswith(".") or (len(toks[-1]) <= 2 and toks[-1].isupper()):
        surname = " ".join(t for t in toks[:-1])
        initial = toks[-1][0]
    else:
        # ESPN-Format: Vorname zuerst
        surname = " ".join(toks[1:])
        initial = toks[0][0]
    return f"{surname} {initial}".lower().replace(".", "").strip()


def norm_player(name):
    if not name:
        return ""
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z ]", "", s.lower()).strip()


SURFACE_DE = {"Clay": "Sand", "Hard": "Hartplatz", "Grass": "Rasen", "Carpet": "Teppich"}


def surface_stats(rows, player):
    """Siege/Niederlagen je Belag ueber die geladenen Jahre."""
    p = player_key(player)
    if not p:
        return {}
    st = {}
    for r in rows:
        surf = r["surface"]
        if not surf:
            continue
        if r["w"] == p:
            st.setdefault(surf, [0, 0])[0] += 1
        elif r["l"] == p:
            st.setdefault(surf, [0, 0])[1] += 1
    out = {}
    for surf, (w, l) in st.items():
        if w + l >= 5:
            out[surf] = {"w": w, "l": l, "pct": round(w / (w + l), 3)}
    return out


def fav_surface(stats):
    if not stats:
        return None
    return max(stats.items(), key=lambda kv: (kv[1]["pct"], kv[1]["w"] + kv[1]["l"]))[0]


def h2h_tennis(rows, p1, p2, n=5):
    """Letzte n direkte Duelle (neueste zuerst)."""
    a, b = player_key(p1), player_key(p2)
    if not a or not b or a == b:
        return []
    meetings = []
    for r in rows:
        if {r["w"], r["l"]} == {a, b}:
            meetings.append({
                "year": r["date"][:4],
                "date": r["date"],
                "tourney": r["tourney"],
                "surface": SURFACE_DE.get(r["surface"], r["surface"]),
                "winner": p1 if r["w"] == a else p2,
                "winnerIsP1": r["w"] == a,
                "score": r["score"],
                "rnd": r["rnd"],
            })
    meetings.sort(key=lambda x: x["date"], reverse=True)
    for m in meetings:
        m.pop("date", None)
    return meetings[:n]


def sack_last_rank(rows, player, max_age_days=400):
    """Zuletzt bekannte Weltranglisten-Position aus den Matchdaten (nicht zu alt)."""
    p = player_key(player)
    cutoff = (NOW - timedelta(days=max_age_days)).strftime("%Y%m%d")
    best = (None, "")
    for r in rows:
        if r["date"] < cutoff:
            continue
        if r["w"] == p and r["wrank"]:
            if r["date"] > best[1]:
                best = (r["wrank"], r["date"])
        elif r["l"] == p and r["lrank"]:
            if r["date"] > best[1]:
                best = (r["lrank"], r["date"])
    try:
        return int(float(best[0])) if best[0] else None
    except ValueError:
        return None


def espn_rankings(tour):
    """athleteId -> rank und name -> rank."""
    idx = http_get(f"https://sports.core.api.espn.com/v2/sports/tennis/leagues/{tour}/rankings")
    by_id, by_name = {}, {}
    if not idx or not idx.get("items"):
        return by_id, by_name
    ref = idx["items"][0].get("$ref", "").replace("http://", "https://")
    doc = http_get(ref)
    if not doc:
        return by_id, by_name
    for entry in doc.get("ranks", []):
        rank = entry.get("current")
        ath_ref = (entry.get("athlete") or {}).get("$ref", "")
        m = re.search(r"/athletes/(\d+)", ath_ref)
        if rank and m:
            by_id[m.group(1)] = rank
        # manche Feeds haben den Namen inline
        nm = (entry.get("athlete") or {}).get("displayName")
        if rank and nm:
            by_name[norm_team(nm)] = rank
    return by_id, by_name


def tennis_text(p1, p2, r1, r2, p1win, surf_de, s1, s2, h2h):
    """Deterministischer deutscher Analysetext fuer ein Tennis-Match."""
    parts = []
    if r1 and r2:
        parts.append(f"{p1} (Weltrangliste #{r1}) trifft auf {p2} (#{r2}).")
    elif r1:
        parts.append(f"{p1} (#{r1}) trifft auf {p2}, der/die aktuell außerhalb der Top-Platzierungen geführt wird.")
    elif r2:
        parts.append(f"{p1} – aktuell ohne Top-Ranking – trifft auf {p2} (#{r2}).")
    else:
        parts.append(f"{p1} trifft auf {p2}; für beide liegt kein aktuelles Top-Ranking vor.")
    if surf_de:
        b1 = s1.get("bilanz")
        b2 = s2.get("bilanz")
        if b1 and b2:
            parts.append(f"Auf {surf_de} liegt die Siegquote von {p1} zuletzt bei "
                         f"{round(b1['pct']*100)} % ({b1['w']}:{b1['l']}), {p2} kommt auf "
                         f"{round(b2['pct']*100)} % ({b2['w']}:{b2['l']}).")
        elif b1:
            parts.append(f"Auf {surf_de} liegt die Siegquote von {p1} zuletzt bei "
                         f"{round(b1['pct']*100)} % ({b1['w']}:{b1['l']}).")
        elif b2:
            parts.append(f"Auf {surf_de} liegt die Siegquote von {p2} zuletzt bei "
                         f"{round(b2['pct']*100)} % ({b2['w']}:{b2['l']}).")
        if s1.get("fav") and s1["fav"] != surf_de:
            parts.append(f"{p1}s stärkster Belag ist eigentlich {s1['fav']}.")
        if s2.get("fav") and s2["fav"] != surf_de:
            parts.append(f"{p2}s stärkster Belag ist eigentlich {s2['fav']}.")
    if h2h:
        w1 = sum(1 for m in h2h if m.get("winnerIsP1"))
        w2 = len(h2h) - w1
        last = h2h[0]
        if len(h2h) == 1:
            score = f" mit {last['score']}" if last.get("score") else ""
            parts.append(f"Das bisher einzige Duell ({last['year']}, {last['tourney']}) "
                         f"gewann {last['winner']}{score}.")
        else:
            leader = p1 if w1 > w2 else p2
            if w1 == w2:
                parts.append(f"Der direkte Vergleich der letzten {len(h2h)} Duelle ist mit {w1}:{w2} ausgeglichen.")
            else:
                parts.append(f"Den direkten Vergleich führt {leader} mit {max(w1,w2)}:{min(w1,w2)} "
                             f"aus den letzten {len(h2h)} Duellen an.")
            if last.get("score"):
                parts.append(f"Zuletzt ({last['year']}, {last['tourney']}) gewann {last['winner']} mit {last['score']}.")
    else:
        parts.append("Ein früheres Aufeinandertreffen auf der Tour ist nicht verzeichnet – es ist das erste direkte Duell.")
    if p1win is not None:
        fav = p1 if p1win >= 0.5 else p2
        parts.append(f"Das Ranglisten-Modell sieht {fav} mit {round(max(p1win, 1-p1win)*100)} % vorn.")
    return " ".join(parts)


def football_text(home, away, m):
    """Deterministischer deutscher Analysetext fuer ein Fussballspiel."""
    parts = []
    pred = m.get("prediction")
    ph, pa2 = m.get("posHome"), m.get("posAway")
    if pred:
        p_h, p_d, p_a = pred["pHome"], pred["pDraw"], pred["pAway"]
        if p_h >= 0.55:
            parts.append(f"{home} geht als klarer Favorit ({round(p_h*100)} %) in das Heimspiel gegen {away}.")
        elif p_a >= 0.55:
            parts.append(f"{away} reist als klarer Favorit ({round(p_a*100)} %) zu {home}.")
        elif abs(p_h - p_a) < 0.08:
            parts.append(f"Zwischen {home} und {away} deutet das Modell auf ein Duell auf Augenhöhe hin "
                         f"({round(p_h*100)} % / {round(p_d*100)} % / {round(p_a*100)} %).")
        elif p_h > p_a:
            parts.append(f"{home} ist gegen {away} leicht favorisiert ({round(p_h*100)} % zu {round(p_a*100)} %).")
        else:
            parts.append(f"{away} ist bei {home} leicht favorisiert ({round(p_a*100)} % zu {round(p_h*100)} %).")
    if ph and pa2:
        parts.append(f"In der Tabelle stehen sich Platz {ph} und Platz {pa2} gegenüber.")
    fh, fa = m.get("formHome") or [], m.get("formAway") or []
    if len(fh) >= 3 and len(fa) >= 3:
        sh, sa = fh.count("S"), fa.count("S")
        if sh - sa >= 2:
            parts.append(f"Die Form spricht für {home}: {sh} Siege aus den letzten {len(fh)} Spielen, "
                         f"{away} holte nur {sa}.")
        elif sa - sh >= 2:
            parts.append(f"Die Form spricht für {away}: {sa} Siege aus den letzten {len(fa)} Spielen, "
                         f"{home} holte nur {sh}.")
    h2h = m.get("h2h") or []
    if h2h:
        parts.append(f"Das letzte Duell ({h2h[0]['date'][-4:]}) endete {h2h[0]['score']} "
                     f"({h2h[0]['home']} – {h2h[0]['away']}).")
    if pred:
        tor = "ein torreiches Spiel" if pred["pOver25"] >= 0.55 else (
            "eher wenige Tore" if pred["pOver25"] <= 0.45 else "eine offene Tor-Ausbeute")
        btts = m.get("pBtts")
        btts_txt = f", beide Teams treffen zu {round(btts*100)} %" if btts is not None else ""
        xg = f"{pred['xgHome']:.1f}:{pred['xgAway']:.1f}".replace(".", ",")
        parts.append(f"Das Modell erwartet {tor} (xG {xg}; "
                     f"über 2,5 Tore: {round(pred['pOver25']*100)} %{btts_txt}).")
    sc_h, sc_a = m.get("scorersHome") or [], m.get("scorersAway") or []
    if sc_h or sc_a:
        bits = []
        if sc_h:
            bits.append(f"bei {home} {sc_h[0]['name']} ({sc_h[0]['goals']} Tore)")
        if sc_a:
            bits.append(f"bei {away} {sc_a[0]['name']} ({sc_a[0]['goals']} Tore)")
        parts.append("Gefährlichste Torschützen: " + " und ".join(bits) + ".")
    v = m.get("value")
    if v:
        q = f"{v['odds']:.2f}".replace(".", ",")
        parts.append(f"Quoten-Hinweis: Das Modell hält Tipp {v['outcome']} (Quote {q}) für "
                     f"{round(v['edge']*100)} Prozentpunkte wahrscheinlicher als der Markt.")
    return " ".join(parts)


def scorers_openligadb(shortcut, year):
    """Top-Torschuetzen je Team aus den OpenLigaDB-Tordaten einer Saison."""
    data = http_get(f"https://api.openligadb.de/getmatchdata/{shortcut}/{year}")
    per_team = {}
    if not data:
        return per_team
    for match in data:
        if not match.get("matchIsFinished"):
            continue
        t1 = match["team1"]["teamName"]
        t2 = match["team2"]["teamName"]
        prev1 = prev2 = 0
        goals = sorted(match.get("goals") or [], key=lambda g: (g.get("goalID") or 0))
        for g in goals:
            s1, s2 = g.get("scoreTeam1") or 0, g.get("scoreTeam2") or 0
            scorer = (g.get("goalGetterName") or "").strip()
            team = t1 if s1 > prev1 else (t2 if s2 > prev2 else None)
            prev1, prev2 = s1, s2
            if not scorer or not team or g.get("isOwnGoal"):
                continue
            key = norm_team(team)
            per_team.setdefault(key, {})
            per_team[key][scorer] = per_team[key].get(scorer, 0) + 1
    return {team: sorted(d.items(), key=lambda kv: -kv[1])[:3] for team, d in per_team.items()}


def espn_league_scorers(slug, season_year):
    """Liga-Torschuetzenliste via ESPN Core-API (Namen per Ref-Fetch, gecacht)."""
    result = []
    for stype in (1, 2):
        url = (f"https://sports.core.api.espn.com/v2/sports/soccer/leagues/{slug}/"
               f"seasons/{season_year}/types/{stype}/leaders")
        doc = http_get(url, retries=2)
        if not doc:
            continue
        cats = doc.get("categories") or []
        goals_cat = next((c for c in cats if c.get("name") in ("goalsLeaders", "goals")), None)
        if not goals_cat:
            continue
        for ld in (goals_cat.get("leaders") or [])[:10]:
            ath_ref = (ld.get("athlete") or {}).get("$ref", "").replace("http://", "https://")
            team_ref = (ld.get("team") or {}).get("$ref", "").replace("http://", "https://")
            ath = _ref_cache_get(ath_ref)
            team = _ref_cache_get(team_ref)
            if not ath:
                continue
            result.append({
                "name": ath.get("displayName", "?"),
                "team": (team or {}).get("displayName", ""),
                "goals": int(ld.get("value") or 0),
            })
        if result:
            break
    return result


_ref_cache = {}


def _ref_cache_get(url):
    if not url:
        return None
    if url not in _ref_cache:
        _ref_cache[url] = http_get(url, retries=2)
    return _ref_cache[url]


def tennis_predict(rank1, rank2):
    """Einfaches Ranglisten-Modell: P(1 gewinnt)."""
    if not rank1 and not rank2:
        return None
    r1 = float(rank1) if rank1 else 250.0
    r2 = float(rank2) if rank2 else 250.0
    e = 0.85
    p1 = r2 ** e / (r1 ** e + r2 ** e)
    return round(p1, 4)


# ----------------------------------------------------------------------------
# Hauptprogramm
# ----------------------------------------------------------------------------

def main():
    season = current_season_year()
    season_start = datetime(season, 7, 1, tzinfo=TZ)
    years = [season - 1, season]
    out = {
        "generatedAt": NOW.isoformat(),
        "windowDays": DAYS_AHEAD,
        "days": [(TODAY + timedelta(days=i)).isoformat() for i in range(DAYS_AHEAD)],
        "football": [],
        "tennis": [],
        "meta": {"oddsApiEnabled": bool(ODDS_API_KEY), "season": f"{season}/{str(season+1)[2:]}"},
    }

    fixture_odds = load_fixture_odds_fdcuk()
    print(f"fixtures.csv: {len(fixture_odds)} Spiele mit Quoten")

    for lg in LEAGUES:
        print(f"== {lg['name']} ==")
        # 1) Ergebnisse (Form + Modell)
        if lg["source"] == "openligadb":
            results = load_results_openligadb(lg["oldb"], years)
        else:
            results = load_results_fdcuk(lg["fdcuk"], years) if lg["fdcuk"] else []
            if len(results) < 50:
                results.extend(load_results_espn(lg["espn"], years))
        print(f"  {len(results)} Ergebnisse geladen")

        strengths, league_avg, home_adv = build_strengths(results)
        if lg.get("cup"):
            # Pokale/Turniere: Tabellenplaetze ergeben keinen Sinn
            table, n_teams = {}, 0
        else:
            table, n_teams = compute_table(results, season_start)
        season_results = [r for r in results if r["date"] >= season_start]

        # 2) Kommende Spiele
        if lg["source"] == "openligadb":
            upcoming, first_future = upcoming_openligadb(lg["oldb"], season)
        else:
            upcoming, first_future = upcoming_espn(lg["espn"])
        print(f"  {len(upcoming)} Spiele in den naechsten {DAYS_AHEAD} Tagen")

        # Nationalteams: deutsche Laendernamen (konsistent in Ergebnissen & Spielen)
        is_nat = lg["id"] in ("wm", "em")
        if is_nat:
            for r in results:
                r["home"], r["away"] = de_team(r["home"]), de_team(r["away"])
            for m in upcoming:
                m["home"], m["away"] = de_team(m["home"]), de_team(m["away"])
            if first_future:
                first_future["home"] = de_team(first_future["home"])
                first_future["away"] = de_team(first_future["away"])
            strengths, league_avg, home_adv = build_strengths(results)

        # 3) Torschuetzen (aktuelle Saison, sonst Vorsaison) - nur wenn Spiele anstehen
        scorers, scorers_period = {}, ""
        if upcoming:
            if lg["source"] == "openligadb":
                scorers = scorers_openligadb(lg["oldb"], season)
                scorers_period = "Saison"
                if not scorers:
                    scorers = scorers_openligadb(lg["oldb"], season - 1)
                    scorers_period = "Vorsaison"
            else:
                league_sc = espn_league_scorers(lg["espn"], season)
                scorers_period = "Saison"
                if not league_sc:
                    league_sc = espn_league_scorers(lg["espn"], season - 1)
                    scorers_period = "Vorsaison"
                for s in league_sc:
                    tname = de_team(s["team"]) if lg["id"] in ("wm", "em") else s["team"]
                    scorers.setdefault(norm_team(tname), []).append((s["name"], s["goals"]))

        def team_scorers(team):
            key = norm_team(team)
            if key in scorers:
                return [{"name": n, "goals": g} for n, g in scorers[key][:3]]
            for k, v in scorers.items():
                if k and team_match(key, k):
                    return [{"name": n, "goals": g} for n, g in v[:3]]
            return []

        matches_out = []
        for m in sorted(upcoming, key=lambda x: x["dt"]):
            pred = predict_match(strengths, league_avg, home_adv, m["home"], m["away"])
            odds = find_odds(fixture_odds, lg["fdcuk"], m["home"], m["away"])
            if not odds and ODDS_API_KEY:
                odds = odds_api_football(ODDS_API_SPORT_KEYS.get(lg["id"], ""), m["home"], m["away"])
            value = None
            if odds and pred:
                imp = implied_probs(odds)
                if imp:
                    diffs = [pred["pHome"] - imp[0], pred["pDraw"] - imp[1], pred["pAway"] - imp[2]]
                    best = max(range(3), key=lambda i: diffs[i])
                    if diffs[best] >= 0.05:
                        value = {"outcome": ["1", "X", "2"][best],
                                 "edge": round(diffs[best], 4),
                                 "modelP": [pred["pHome"], pred["pDraw"], pred["pAway"]][best],
                                 "odds": [odds["h"], odds["d"], odds["a"]][best]}
            m_out = {
                "kickoff": m["dt"].isoformat(),
                "home": m["home"], "away": m["away"],
                "homeIcon": m.get("homeIcon"), "awayIcon": m.get("awayIcon"),
                "matchday": m.get("matchday"),
                "formHome": team_form(results, m["home"]),
                "formAway": team_form(results, m["away"]),
                "posHome": table_pos(table, m["home"]),
                "posAway": table_pos(table, m["away"]),
                "nTeams": n_teams if n_teams else None,
                "h2h": head_to_head(results, m["home"], m["away"]),
                "prediction": pred,
                "pBtts": pred["pBtts"] if pred else None,
                "odds": odds,
                "value": value,
                "scorersHome": team_scorers(m["home"]),
                "scorersAway": team_scorers(m["away"]),
                "scorersPeriod": scorers_period,
            }
            m_out["analysis"] = football_text(m["home"], m["away"], m_out)
            matches_out.append(m_out)

        league_out = {
            "id": lg["id"], "name": lg["name"], "flag": lg["flag"],
            "matches": matches_out,
            "seasonHasStarted": len(season_results) > 0,
            "isCup": bool(lg.get("cup")),
        }
        if not matches_out and first_future:
            league_out["nextMatch"] = {
                "kickoff": first_future["dt"].isoformat(),
                "home": first_future["home"], "away": first_future["away"],
                "matchday": first_future.get("matchday"),
            }
        out["football"].append(league_out)

    # ---- Tennis ----
    print("== Tennis ==")
    rank_atp_id, rank_atp_nm = espn_rankings("atp")
    rank_wta_id, rank_wta_nm = espn_rankings("wta")
    print(f"  Rankings: ATP {len(rank_atp_id) or len(rank_atp_nm)}, WTA {len(rank_wta_id) or len(rank_wta_nm)}")
    out["meta"]["rankCounts"] = {"atp": len(rank_atp_id), "wta": len(rank_wta_id)}

    # Historie (Sackmann) fuer Belag-Bilanzen, H2H und Rank-Fallback
    sack = {"ATP": load_sackmann("atp"), "WTA": load_sackmann("wta")}
    print(f"  Historie: ATP {len(sack['ATP'])} Matches, WTA {len(sack['WTA'])} Matches")
    out["meta"]["sackCounts"] = {"atp": len(sack["ATP"]), "wta": len(sack["WTA"])}
    # Turnier -> Belag: Orte und Turniernamen getrennt (neueste Eintraege gewinnen,
    # da Zeilen chronologisch eingelesen werden)
    loc_map, name_map = {}, {}
    for tour_key, rows in sack.items():
        lm, nm2 = {}, {}
        for r in rows:
            if not r["surface"]:
                continue
            lk = norm_player(r.get("loc", ""))
            nk = norm_player(r["tourney"])
            if lk:
                lm[lk] = r["surface"]
            if nk:
                nm2[nk] = r["surface"]
        loc_map[tour_key] = lm
        name_map[tour_key] = nm2

    def tournament_surface(tour_key, name):
        t = norm_player(name)
        tours = (tour_key, "ATP" if tour_key == "WTA" else "WTA")
        # 1) Ort als ganzes Wort im ESPN-Turniernamen (z.B. "gstaad", "umag")
        for tk in tours:
            hits = [(loc, surf) for loc, surf in loc_map.get(tk, {}).items()
                    if len(loc) >= 4 and re.search(rf"(?<![a-z]){re.escape(loc)}(?![a-z])", t)]
            if hits:
                return max(hits, key=lambda x: len(x[0]))[1]
        # 2) Turniername exakt oder als Teilstring (laengster Treffer)
        for tk in tours:
            nm2 = name_map.get(tk, {})
            if t in nm2:
                return nm2[t]
            hits = [(tn, surf) for tn, surf in nm2.items()
                    if len(tn) >= 8 and (tn in t or t in tn)]
            if hits:
                return max(hits, key=lambda x: len(x[0]))[1]
        return None

    ROUND_DE = {
        "round 1": "1. Runde", "round 2": "2. Runde", "round 3": "3. Runde",
        "round 4": "4. Runde", "round of 128": "1. Runde", "round of 64": "2. Runde",
        "round of 32": "Runde der 32", "round of 16": "Achtelfinale",
        "quarterfinal": "Viertelfinale", "quarterfinals": "Viertelfinale",
        "semifinal": "Halbfinale", "semifinals": "Halbfinale",
        "final": "Finale", "finals": "Finale",
    }

    seen = set()
    tennis_out = []
    for i in range(DAYS_AHEAD):
        day = TODAY + timedelta(days=i)
        day_matches = []
        for tour in ("atp", "wta"):
            day_matches.extend(espn_tennis_day(tour, day))
        # TheSportsDB nur als Ergaenzung, wenn ESPN fuer diese Tour an dem Tag
        # nichts liefert. Community-Daten ordnen Turniere teils falsch zu,
        # daher werden solche Eintraege als "unbestaetigt" markiert.
        espn_tours_today = {m["tour"] for m in day_matches}
        for m in tsdb_tennis_day(day):
            if m["tour"] not in espn_tours_today:
                day_matches.append(m)
        for m in day_matches:
            # Dedupe OHNE Tour: kombinierte Turniere tauchen in beiden Feeds auf
            key = (norm_team(m["p1"]["name"]), norm_team(m["p2"]["name"]), m["dt"].date().isoformat())
            key_rev = (key[1], key[0], key[2])
            if key in seen or key_rev in seen:
                continue
            seen.add(key)
            by_id = rank_atp_id if m["tour"] == "ATP" else rank_wta_id
            by_nm = rank_atp_nm if m["tour"] == "ATP" else rank_wta_nm
            rows = sack.get(m["tour"], [])
            n1, n2 = m["p1"]["name"], m["p2"]["name"]
            # Ranking: Weltrangliste (Top 150), sonst letzter bekannter Rang aus der Historie
            r1 = (by_id.get(m["p1"]["id"]) or by_nm.get(norm_team(n1))
                  or sack_last_rank(rows, n1))
            r2 = (by_id.get(m["p2"]["id"]) or by_nm.get(norm_team(n2))
                  or sack_last_rank(rows, n2))
            p1win = tennis_predict(r1, r2)
            rnd = (m.get("round") or "").strip()
            rnd = ROUND_DE.get(rnd.lower(), rnd)
            # Belag & Bilanzen
            surf = tournament_surface(m["tour"], m["tournament"])
            surf_de = SURFACE_DE.get(surf, surf) if surf else None
            st1, st2 = surface_stats(rows, n1), surface_stats(rows, n2)
            s1 = {"bilanz": st1.get(surf), "fav": SURFACE_DE.get(fav_surface(st1), fav_surface(st1)),
                  "alle": {SURFACE_DE.get(k, k): v for k, v in st1.items()}}
            s2 = {"bilanz": st2.get(surf), "fav": SURFACE_DE.get(fav_surface(st2), fav_surface(st2)),
                  "alle": {SURFACE_DE.get(k, k): v for k, v in st2.items()}}
            h2h = h2h_tennis(rows, n1, n2)
            tennis_out.append({
                "start": m["dt"].isoformat(),
                "tour": m["tour"], "tournament": m["tournament"],
                "round": rnd,
                "surface": surf_de,
                "p1": {"name": n1, "rank": r1,
                       "onSurface": s1["bilanz"], "favSurface": s1["fav"]},
                "p2": {"name": n2, "rank": r2,
                       "onSurface": s2["bilanz"], "favSurface": s2["fav"]},
                "pP1": p1win,
                "h2h": h2h,
                "analysis": tennis_text(n1, n2, r1, r2, p1win, surf_de, s1, s2, h2h),
                "unconfirmed": m.get("src") == "tsdb",
                "timeTBD": bool(m.get("timeTBD")),
            })
    tennis_out.sort(key=lambda x: x["start"])
    out["tennis"] = tennis_out
    print(f"  {len(tennis_out)} Tennis-Matches in den naechsten {DAYS_AHEAD} Tagen")

    os.makedirs(os.path.join(os.path.dirname(__file__), "..", "data"), exist_ok=True)
    path = os.path.join(os.path.dirname(__file__), "..", "data", "data.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1, default=str)
    print(f"OK -> {path}")


if __name__ == "__main__":
    main()
